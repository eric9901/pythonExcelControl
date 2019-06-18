from openpyxl import load_workbook       #load the packect for I/O excelfile

def loadXFile(fileName):
    wb = load_workbook(fileName+'.xlsx')     #load the excel file
    return wb;

def loadSheet(sheetname):                 #有local variable 係def , 係主程式係global
    if(sheetname=="active"):
        a_sheet=wb.active
    else:
        a_sheet=wb[sheetname]
    return a_sheet

def showSheetName():
    print(wb.sheetnames)

def setBox(row, column=None):
    if (column is None):
        box = sheet[str(row)]
    else:
        box = sheet.cell(row=row, column=column)
    return box

def printAllvalue(order):
    if(order=="row"):
        print('----------------------------')  # print cell value by row
        for row in sheet.rows:
            for cell in row:
                print(cell.value)
    elif (order=="column"):
        print('--------------------')  # print cell value by column

        for column in sheet.columns:
            for cell in column:
                print(cell.value)
    else:
        print('wrong command')



def getBoxValue(box):
    return box.value

#main coding
wb=loadXFile('testcase1')
showSheetName()             #get the sheet name that are reading

a_sheet= wb['工作表1']#get the sheet that by the sheet name
sheet=loadSheet('active')                          #get the sheet that are reading
b4=setBox('B4')

#print(f'({b4.column}, {b4.row}) is {b4.value}') # f係f-string 即Literal String Interpolatio, 可以用大括號係同一''內用var name 表示var value
b4Too= setBox(3,2)            #直接用row,column 數揾個位
print(b4.value)
print(getBoxValue(b4Too))
printAllvalue("column")






