from openpyxl import load_workbook       #load the packect for I/O excelfile

wb = load_workbook('testcase1.xlsx')     #load the excel file
print(wb.get_sheet_names())              #get the sheet name that are reading
a_sheet= wb.get_sheet_by_name('工作表1') #get the sheet that by the sheet name
print(a_sheet)
sheet=wb.active                          #get the sheet that are reading
b4=sheet['B4']
print(f'({b4.column}, {b4.row}) is {b4.value}') # f係f-string 即Literal String Interpolatio, 可以用大括號係同一''內用var name 表示var value
b4_too = sheet.cell(row=4, column=2)            #直接用row,column 數揾個位
print(b4_too.value)
